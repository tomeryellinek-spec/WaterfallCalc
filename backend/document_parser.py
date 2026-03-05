import os
import json
import io
from dotenv import load_dotenv
from anthropic import AsyncAnthropic
from openpyxl import load_workbook

load_dotenv()
client = AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))

SYSTEM_PROMPT = """You are a financial document parser. Extract ALL structured data from the provided document.

Return ONLY valid JSON matching the schema below. Do not include any text outside the JSON.

Schema:
{
  "seniorDebts": [{"name": str, "holderName": str, "principal": number, "accruedInterest": number, "interestRate": number, "seniority": number}],
  "safes": [{"investorName": str, "investmentAmount": number, "valuationCap": number, "discountRate": number}],
  "preferredShares": [{"className": str, "pricePerShare": number, "liquidationMultiple": number, "participation": "non-participating"|"participating"|"capped", "participationCap": number, "seniority": number, "conversionRatio": number, "holders": [{"name": str, "shares": number}]}],
  "commonShares": [{"className": str, "holders": [{"name": str, "shares": number}]}],
  "options": [{"holderName": str, "shares": number, "strikePrice": number, "vested": number}],
  "warrants": [{"holderName": str, "shares": number, "exercisePrice": number}]
}

CRITICAL RULES FOR COMPLETENESS:
- You MUST extract EVERY single holder/investor/shareholder listed in the document. Do NOT skip or summarize.
- For cap tables: extract ALL share classes (Common, Series Seed, Series A, Series B, etc.) and ALL holders within each class.
- If a holder appears in multiple series/classes, include them in each relevant class's holders array.
- Count carefully: if the document lists 20 shareholders, your output must have 20 shareholders. Do not truncate.
- Each preferred share class (Series Seed, Series A, Series A-1, Series B, etc.) must be a separate entry in preferredShares with its own holders array.

Other rules:
- Only include fields that are present in the document. Omit arrays that have no data.
- For interest rates, use decimal form (5% = 0.05).
- For discount rates on SAFEs, use decimal form (20% = 0.20).
- For vested percentage on options, use decimal form (100% = 1.0).
- Seniority: 1 = most senior. Higher numbers = more junior. Assign seniority by series order (later series = more senior).
- If liquidation multiple is not specified, default to 1.
- If participation type is not specified, default to "non-participating".
- conversionRatio defaults to 1 if not specified.
"""

DOC_TYPE_HINTS = {
    "cap_table": "This document is a capitalization table (cap table). Extract EVERY share class and EVERY holder/shareholder. Do not skip any rows. Include all series (Seed, A, B, etc.) and all common shareholders, option holders, and warrant holders.",
    "loan_agreement": "This document is a loan/debt agreement. Focus on extracting debt terms: principal, interest, lender, security/seniority.",
    "safe": "This document is a SAFE (Simple Agreement for Future Equity). Focus on extracting investor, investment amount, valuation cap, and discount rate.",
    "warrant": "This document is a warrant agreement. Focus on extracting holder, number of shares, exercise price.",
    "option_plan": "This document is a stock option plan or grant. Focus on extracting holders, shares, strike prices, and vesting.",
    "other": "Extract any financial instrument data you can identify from this document.",
}


def parse_excel(content: bytes, ext: str) -> str:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        # Retry without formatting (handles invalid color/style XML)
        import zipfile
        import xml.etree.ElementTree as ET

        lines = []
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    sheet_num = name.replace("xl/worksheets/sheet", "").replace(".xml", "")
                    lines.append(f"=== Sheet {sheet_num} ===")
                    tree = ET.parse(zf.open(name))
                    ns = {"": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                    # Try to load shared strings for text cells
                    strings = []
                    if "xl/sharedStrings.xml" in zf.namelist():
                        st_tree = ET.parse(zf.open("xl/sharedStrings.xml"))
                        for si in st_tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"):
                            texts = si.itertext()
                            strings.append("".join(texts))
                    for row_el in tree.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"):
                        vals = []
                        for cell in row_el.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
                            v_el = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
                            val = v_el.text if v_el is not None else ""
                            if cell.get("t") == "s" and val and strings:
                                idx = int(val)
                                val = strings[idx] if idx < len(strings) else val
                            vals.append(val or "")
                        lines.append("\t".join(vals))
                    lines.append("")
        if not lines:
            raise ValueError("Could not read Excel file")
        return "\n".join(lines)

    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(vals))
        lines.append("")
    return "\n".join(lines)


def parse_cap_table_excel(content: bytes) -> dict | None:
    """Directly parse a cap table Excel file by reading the tab-delimited text from parse_excel."""
    import re

    text = parse_excel(content, ".xlsx")
    if not text:
        return None

    # Parse into rows
    lines = text.split("\n")
    rows = []
    for line in lines:
        if line.startswith("=== Sheet"):
            continue
        if not line.strip():
            continue
        rows.append(line.split("\t"))

    if len(rows) < 3:
        return None

    # Find header row - look for row with class names
    header_row = None
    header_idx = None
    for i, row in enumerate(rows):
        joined = " ".join(row).lower()
        if ("common" in joined and ("preferred" in joined or "series" in joined or "seed" in joined)) or \
           ("shareholder" in joined and ("series" in joined or "seed" in joined or "preferred" in joined)):
            header_row = [c.strip() for c in row]
            header_idx = i
            break

    if header_row is None or header_idx is None:
        return None

    # Check sub-header
    sub_header = None
    if header_idx + 1 < len(rows):
        sub_header = [c.strip() for c in rows[header_idx + 1]]

    # Map columns
    col_map = {}
    shareholder_col = 0
    warrant_col = None
    option_col = None

    for ci, val in enumerate(header_row):
        vl = val.lower()
        if "shareholder" in vl or (ci == 0 and vl != ""):
            shareholder_col = ci
        elif "common" in vl:
            col_map[ci] = {"type": "common", "class_name": val.replace(" shares", "").replace(" Shares", "").strip()}
        elif "option" in vl:
            option_col = ci
        elif any(kw in vl for kw in ["series", "seed", "preferred"]):
            class_name = val.replace(" Shares", "").replace(" shares", "").strip()
            col_map[ci] = {"type": "preferred", "class_name": class_name}

    # Check sub-header for Shares/Warrants/Options columns
    if sub_header:
        for ci, sv in enumerate(sub_header):
            svl = sv.lower().strip()
            if svl == "warrants":
                warrant_col = ci
            elif svl == "options" or "option" in svl:
                option_col = ci
            elif svl == "shares" and ci not in col_map:
                # Common shares column (usually right after shareholder name)
                if ci == shareholder_col + 1 or (header_row[ci].strip() == "" and ci <= 2):
                    col_map[ci] = {"type": "common", "class_name": "Common"}

    # Parse data rows
    preferred_data = {}
    common_data = {}
    options_data = []
    warrants_data = []

    data_start = header_idx + 1
    if sub_header and any(s.strip().lower() in ("shares", "warrants", "options", "total holdings (%)", "fully diluted (%)") for s in sub_header):
        data_start = header_idx + 2

    for row in rows[data_start:]:
        name = row[shareholder_col].strip() if shareholder_col < len(row) else ""
        if not name:
            continue
        nl = name.lower()
        if any(kw in nl for kw in ["total", "unallocated", "outstanding option", "exercised"]):
            continue

        for ci, info in col_map.items():
            if ci >= len(row) or not row[ci].strip():
                continue
            try:
                shares = float(row[ci].strip())
            except (ValueError, TypeError):
                continue
            if shares <= 0:
                continue

            if info["type"] == "preferred":
                preferred_data.setdefault(info["class_name"], []).append({"name": name, "shares": shares})
            elif info["type"] == "common":
                common_data.setdefault(info["class_name"], []).append({"name": name, "shares": shares})

        if warrant_col is not None and warrant_col < len(row) and row[warrant_col].strip():
            try:
                shares = float(row[warrant_col].strip())
                if shares > 0:
                    warrants_data.append({"holderName": name, "shares": shares, "exercisePrice": 0})
            except (ValueError, TypeError):
                pass

        if option_col is not None and option_col < len(row) and row[option_col].strip():
            try:
                shares = float(row[option_col].strip())
                if shares > 0:
                    options_data.append({"holderName": name, "shares": shares, "strikePrice": 0, "vested": 1.0})
            except (ValueError, TypeError):
                pass

    if not preferred_data and not common_data:
        return None

    result = {}

    if preferred_data:
        def series_order(name: str) -> int:
            nl = name.lower()
            if "seed" in nl:
                # Sub-sort seed variants
                if "3" in nl: return 1
                if "2" in nl: return 2
                if "1" in nl: return 3
                return 2
            match = re.search(r'(?:preferred\s+|series\s+)?([a-c])[-]?(\d)?', nl)
            if match:
                letter = match.group(1)
                sub = int(match.group(2)) if match.group(2) else 0
                return (ord(letter) - ord('a') + 4) * 10 + sub
            return 100

        preferred_list = []
        sorted_classes = sorted(preferred_data.items(), key=lambda x: series_order(x[0]))
        num_classes = len(sorted_classes)
        for idx, (class_name, holders) in enumerate(sorted_classes):
            # Later series = more senior = lower seniority number (paid first)
            preferred_list.append({
                "className": class_name,
                "pricePerShare": 0,
                "liquidationMultiple": 1,
                "participation": "non-participating",
                "participationCap": 0,
                "seniority": num_classes - idx,
                "conversionRatio": 1,
                "holders": holders,
            })
        result["preferredShares"] = preferred_list

    if common_data:
        result["commonShares"] = [{"className": cn, "holders": h} for cn, h in common_data.items()]

    if options_data:
        result["options"] = options_data
    if warrants_data:
        result["warrants"] = warrants_data

    return result


async def parse_with_claude(
    content: str,
    document_type: str,
    is_text: bool = True,
    media_type: str | None = None,
) -> dict:
    hint = DOC_TYPE_HINTS.get(document_type, DOC_TYPE_HINTS["other"])

    if is_text:
        # Log the raw text being sent for debugging
        with open("last_parsed_input.txt", "w", encoding="utf-8") as f:
            f.write(content)
        print(f"[DEBUG] Sending {len(content)} chars to Claude (saved to last_parsed_input.txt)")
        user_content = [
            {"type": "text", "text": f"{hint}\n\nDocument content:\n\n{content}"}
        ]
    else:
        user_content = [
            {
                "type": "document",
                "source": {"type": "base64", "media_type": media_type, "data": content},
            },
            {"type": "text", "text": hint},
        ]

    response = await client.messages.create(
        model="claude-opus-4-6",
        max_tokens=16384,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_content}],
    )

    raw_text = response.content[0].text if response.content else "{}"

    # Extract JSON from response
    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError:
        # Try to find JSON block in the response
        start = raw_text.find("{")
        end = raw_text.rfind("}") + 1
        if start >= 0 and end > start:
            data = json.loads(raw_text[start:end])
        else:
            data = {}

    # Log Claude's response for debugging
    with open("last_parsed_output.json", "w", encoding="utf-8") as f:
        f.write(raw_text)
    print(f"[DEBUG] Claude response: {len(raw_text)} chars (saved to last_parsed_output.json)")

    return {
        "documentType": document_type,
        "data": data,
        "rawText": raw_text[:2000],
    }
